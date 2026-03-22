import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import random
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import os

# -------------------------------
# Page Config
# -------------------------------
st.set_page_config(page_title="Workout Tracker Pro", layout="wide")

# -------------------------------
# LIGHT UI Styling
# -------------------------------
st.markdown("""
<style>
body {
    background-color: #f7f9fc;
}
.block-container {
    padding-top: 1rem;
}
.card {
    background: white;
    padding: 20px;
    border-radius: 12px;
    box-shadow: 0px 2px 10px rgba(0,0,0,0.1);
}
.stButton>button {
    background-color: #2e7dff;
    color: white;
    border-radius: 8px;
    height: 3em;
    width: 100%;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

# -------------------------------
# Motivational Quotes
# -------------------------------
quotes = [
    "Fitness is not quitting. It’s getting up one more time 💪",
    "Discipline creates results 🔥",
    "Push beyond limits 🚀",
    "Consistency beats everything 💯",
    "Your only competition is you 👊"
]

quote = random.choice(quotes)


# -------------------------------
# HEADER TEXT
# -------------------------------
st.markdown(f"""
<h1 style='text-align:center;'>💪 Workout Tracker Pro</h1>
<h4 style='text-align:center; color:gray;'>{quote}</h4>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# -------------------------------
# Layout
# -------------------------------
col1, col2 = st.columns([2, 1])

# -------------------------------
# INPUT CARD
# -------------------------------
with col1:
    st.markdown('<div class="card">', unsafe_allow_html=True)

    st.subheader("🏋️ Workout Input")

    workout = st.text_input("Workout Name", placeholder="Bench Press")
    date = st.date_input("Date")

    sets = st.number_input("Number of Sets", min_value=1, max_value=10, value=3)

    st.markdown("### Enter details for each set")

    set_data = []

    for i in range(sets):
        c1, c2 = st.columns(2)

        with c1:
            weight = st.number_input(f"Set {i+1} - Weight (kg)", min_value=0, value=20, key=f"w_{i}")
        with c2:
            reps = st.number_input(f"Set {i+1} - Reps", min_value=1, value=10, key=f"r_{i}")

        set_data.append((weight, reps))

    generate = st.button("🚀 Generate Workout Analysis")

    st.markdown('</div>', unsafe_allow_html=True)

# -------------------------------
# SIDE PANEL
# -------------------------------
with col2:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("📌 Tips")
    st.write("""
    • Progressive overload  
    • Maintain proper form  
    • Rest between sets  
    """)
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("🔥 Motivation")
    st.success("Small daily progress = Big results 💪")
    st.markdown('</div>', unsafe_allow_html=True)

# -------------------------------
# Excel Function (with borders)
# -------------------------------
def create_excel(df):
    wb = Workbook()
    ws = wb.active

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Header
    for col_num, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col)
        cell.border = thin
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_num, row in enumerate(df.values, 2):
        for col_num, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# -------------------------------
# OUTPUT
# -------------------------------
if generate:
    if not workout:
        st.warning("Please enter workout name")
    else:
        total_volume = 0
        data = []
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for i, (w, r) in enumerate(set_data):
            volume = w * r
            total_volume += volume

            data.append({
                "Workout": workout,
                "Date": date.strftime("%Y-%m-%d"),
                "DateTime": current_time,
                "Set": i + 1,
                "Weight (kg)": w,
                "Reps": r,
                "Volume": volume
            })

        df = pd.DataFrame(data)

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)

        st.subheader("📊 Workout Breakdown")
        st.dataframe(df, use_container_width=True)

        st.success(f"🔥 Total Volume: {total_volume} kg")

        excel = create_excel(df)
        file_name = f"workout_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            "⬇️ Download Excel",
            excel,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown('</div>', unsafe_allow_html=True)